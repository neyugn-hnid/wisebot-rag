import logging
import os
from io import BytesIO

from fastapi import FastAPI, File, HTTPException, UploadFile
from kreuzberg import ExtractionConfig, extract_bytes

from app.chunking import chunk_text

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="kreuzberg-service", version="1.0.0")

EXTRACT_CONFIG = ExtractionConfig(use_cache=True)
DEFAULT_CHUNK_SIZE = int(os.getenv("DEFAULT_CHUNK_SIZE", "800"))
DEFAULT_CHUNK_OVERLAP = int(os.getenv("DEFAULT_CHUNK_OVERLAP", "150"))

# MIME types cho spreadsheet
EXCEL_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
}


def _xlsx_to_markdown(file_bytes: bytes, filename: str) -> tuple[str, list[dict]]:
    """Doc truc tiep xlsx bang openpyxl, convert moi dong -> markdown card."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]

    chunks: list[dict] = []
    full_text_parts: list[str] = []
    idx = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        # Tim ten san pham (cot thu 2 hoac cot co "ten")
        title = ""
        for key in ("ten_san_pham", "ten", "name", "san_pham", "title"):
            val = data.get(key)
            if val and str(val).strip():
                title = str(val).strip()
                break
        if not title:
            title = str(list(data.values())[1]) if len(data) > 1 else str(list(data.values())[0])

        parts = [f"## {title}", ""]
        for h, v in data.items():
            if v is None or str(v).strip() == "":
                continue
            label = h.replace("_", " ").title()
            parts.append(f"- **{label}**: {v}")
        parts.append("")

        card = "\n".join(parts)
        full_text_parts.append(card)
        chunks.append({"index": idx, "content": card})
        idx += 1

    wb.close()
    full_text = "\n".join(full_text_parts)
    logger.info("XLSX parsed: %s -> %d rows, %d chunks", filename, idx, idx)
    return full_text, chunks


def _build_chunks(text: str, raw_chunks: object | None, format_type: str = "") -> list[dict]:
    chunk_sources: list[str] = []

    if isinstance(raw_chunks, list):
        for chunk in raw_chunks:
            content = chunk.get("content") if isinstance(chunk, dict) else getattr(chunk, "content", None)
            if isinstance(content, str) and content.strip():
                chunk_sources.append(content)

    if not chunk_sources:
        chunk_sources = [text]

    chunks: list[str] = []
    for source in chunk_sources:
        source_chunks = chunk_text(
            source,
            chunk_size=DEFAULT_CHUNK_SIZE,
            chunk_overlap=DEFAULT_CHUNK_OVERLAP,
        )
        chunks.extend(source_chunks or [source.strip()])

    return [{"index": i, "content": content} for i, content in enumerate(chunks) if content.strip()]


@app.get("/health")
async def healthcheck() -> dict:
    return {"status": "ok", "service": "kreuzberg-service"}


@app.post("/parse")
async def parse(file: UploadFile = File(...)) -> dict:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    mime = file.content_type or ""

    # ── XLSX/Excel: doc truc tiep bang openpyxl -> markdown ───────────
    if mime in EXCEL_MIMES or (file.filename or "").lower().endswith(".xlsx"):
        try:
            text, chunks = _xlsx_to_markdown(content, file.filename or "spreadsheet.xlsx")
            return {
                "filename": file.filename,
                "text": text,
                "chunks": chunks,
                "chars": len(text),
                "mime_type": "excel",
            }
        except Exception as exc:
            logger.exception("openpyxl failed for %s, falling back to kreuzberg", file.filename)

    # ── Cac dinh dang khac: dung kreuzberg ────────────────────────────
    try:
        result = await extract_bytes(content, mime_type=mime or "application/octet-stream", config=EXTRACT_CONFIG)
        if not result.content:
            raise HTTPException(status_code=422, detail="No content extracted")

        format_type = str(result.metadata.get("format_type", "") if isinstance(result.metadata, dict) else "")
        chunks = _build_chunks(result.content, getattr(result, "chunks", None), format_type)

        logger.info("Parsed %s: %d chars, %d chunks, mime=%s",
                    file.filename, len(result.content), len(chunks), format_type or "unknown")

        return {
            "filename": file.filename,
            "text": result.content,
            "chunks": chunks,
            "chars": len(result.content),
            "mime_type": format_type or None,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Kreuzberg parse failed for %s", file.filename)
        raise HTTPException(status_code=500, detail=f"Kreuzberg parse failed: {exc}") from exc
